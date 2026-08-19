import os
import shutil
import sys
from datetime import datetime
from typing import List, Dict, Any, Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from tests.config import EXCEL_TRACKER_PATH

class TestReporter:
    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super(TestReporter, cls).__new__(cls)
            cls._instance.results = []
            cls._instance.run_timestamp = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
        return cls._instance

    def reset(self):
        self.results.clear()
        self.run_timestamp = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")

    def record(
        self,
        test_id: str,
        category: str,
        test_case: str,
        priority: str,
        expected_result: str,
        actual_result: str,
        status: str,
        bug_id: str = "",
        notes: str = ""
    ):
        emoji = "PASS" if status == "PASSED" else ("GAP" if status == "GAP" else "FAIL")
        short_actual = str(actual_result)[:85]
        safe_output = f"  [{emoji}] [{test_id}] {short_actual}".encode(sys.stdout.encoding or "utf-8", errors="replace").decode(sys.stdout.encoding or "utf-8")
        try:
            print(safe_output)
        except Exception:
            print(f"  [{emoji}] [{test_id}] {short_actual.encode('ascii', errors='replace').decode('ascii')}")

        self.results.append({
            "test_id": test_id,
            "category": category,
            "test_case": test_case,
            "priority": priority,
            "expected_result": expected_result,
            "actual_result": str(actual_result),
            "status": status,
            "bug_id": bug_id,
            "notes": notes,
            "timestamp": self.run_timestamp
        })

    def print_summary(self):
        total = len(self.results)
        passed = sum(1 for r in self.results if r["status"] == "PASSED")
        failed = sum(1 for r in self.results if r["status"] == "FAILED")
        gaps = sum(1 for r in self.results if r["status"] == "GAP")

        print("\n" + "=" * 80)
        print(f"   LIVE TEST SUITE SUMMARY - {self.run_timestamp}")
        print(f"   TOTAL: {total}  |  PASSED: {passed}  |  FAILED: {failed}  |  GAPS: {gaps}")
        if total > 0:
            pass_rate = (passed / total) * 100
            print(f"   PASS RATE: {pass_rate:.1f}%")
        print("=" * 80 + "\n")

    def export_to_excel(self, excel_path: str = EXCEL_TRACKER_PATH):
        if not os.path.exists(excel_path):
            print(f"Warning: Excel file not found at {excel_path}. Creating new workbook.")
            wb = openpyxl.Workbook()
        else:
            wb = openpyxl.load_workbook(excel_path)

        sheet_name = "Live Runtime Results"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        header_bg = "1F4E79"
        header_fg = "FFFFFF"
        pass_bg = "D4EDDA"
        pass_fg = "155724"
        fail_bg = "F8D7DA"
        fail_fg = "721C24"
        gap_bg = "FFF3CD"
        gap_fg = "856404"

        headers = [
            "Test ID", "Category", "Test Case", "Priority", "Expected Result",
            "Actual Result", "Status", "Bug / Issue ID", "Evidence / Notes", "Timestamp"
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = PatternFill(start_color=header_bg, end_color=header_bg, fill_type="solid")
            cell.font = Font(name="Calibri", size=10, bold=True, color=header_fg)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_idx, r in enumerate(self.results, 2):
            ws.cell(row=row_idx, column=1, value=r["test_id"])
            ws.cell(row=row_idx, column=2, value=r["category"])
            ws.cell(row=row_idx, column=3, value=r["test_case"])
            ws.cell(row=row_idx, column=4, value=r["priority"])
            ws.cell(row=row_idx, column=5, value=r["expected_result"])
            ws.cell(row=row_idx, column=6, value=r["actual_result"])

            status_cell = ws.cell(row=row_idx, column=7, value=r["status"])
            if r["status"] == "PASSED":
                status_cell.fill = PatternFill(start_color=pass_bg, end_color=pass_bg, fill_type="solid")
                status_cell.font = Font(name="Calibri", size=10, bold=True, color=pass_fg)
            elif r["status"] == "FAILED":
                status_cell.fill = PatternFill(start_color=fail_bg, end_color=fail_bg, fill_type="solid")
                status_cell.font = Font(name="Calibri", size=10, bold=True, color=fail_fg)
            else:
                status_cell.fill = PatternFill(start_color=gap_bg, end_color=gap_bg, fill_type="solid")
                status_cell.font = Font(name="Calibri", size=10, bold=True, color=gap_fg)

            ws.cell(row=row_idx, column=8, value=r["bug_id"])
            ws.cell(row=row_idx, column=9, value=r["notes"])
            ws.cell(row=row_idx, column=10, value=r["timestamp"])

        column_widths = {
            "A": 18, "B": 35, "C": 60, "D": 10, "E": 40,
            "F": 60, "G": 12, "H": 15, "I": 40, "J": 20
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        ws.auto_filter.ref = ws.dimensions

        try:
            if "Executive Summary" in wb.sheetnames:
                ws_sum = wb["Executive Summary"]
                ws_sum["B3"] = self.run_timestamp
                ws_sum["B4"] = len(self.results)
                ws_sum["B5"] = sum(1 for r in self.results if r["status"] == "PASSED")
                ws_sum["B6"] = sum(1 for r in self.results if r["status"] == "FAILED")
                ws_sum["B7"] = sum(1 for r in self.results if r["status"] == "GAP")
        except Exception as e:
            print(f"Notice: Could not update Executive Summary: {e}")

        temp_path = excel_path.replace(".xlsx", "_temp.xlsx")
        wb.save(temp_path)
        shutil.move(temp_path, excel_path)
        print(f"Results exported successfully to: {excel_path}")

reporter = TestReporter()
